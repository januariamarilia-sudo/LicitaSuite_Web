from pathlib import Path
import re
import unicodedata

try:
    import openpyxl
except Exception:
    openpyxl = None


def only_digits(value):
    return re.sub(r"\D+", "", "" if value is None else str(value))


def normalize_text(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.upper().split())


def clean_cell(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".0", "").isdigit():
        text = text.replace(".0", "")
    return " ".join(text.split()).strip()


def split_address_city_uf(endereco):
    """
    Lê endereço de mala direta, por exemplo:
    Rodovia JK 459, KM 99 S/N Galpão, Bairro Santa Edwirges, no Município de Pouso Alegre - MG

    Retorna:
    endereco_limpo, municipio, uf
    """
    text = clean_cell(endereco)
    if not text:
        return "", "", ""

    patterns = [
        r"^(?P<addr>.*?),?\s+no\s+Munic[ií]pio\s+de\s+(?P<city>.+?)\s*[-/]\s*(?P<uf>[A-Z]{2})\.?$",
        r"^(?P<addr>.*?),?\s+Munic[ií]pio\s+de\s+(?P<city>.+?)\s*[-/]\s*(?P<uf>[A-Z]{2})\.?$",
        r"^(?P<addr>.*?),?\s+em\s+(?P<city>.+?)\s*[-/]\s*(?P<uf>[A-Z]{2})\.?$",
    ]

    for pattern in patterns:
        m = re.search(pattern, text, flags=re.IGNORECASE)
        if m:
            addr = clean_cell(m.group("addr")).rstrip(",")
            city = clean_cell(m.group("city"))
            uf = clean_cell(m.group("uf")).upper()
            return addr, city, uf

    # fallback: tenta final "Cidade - UF"
    m = re.search(r"(?P<city>[A-Za-zÀ-ÿ\s]+)\s*[-/]\s*(?P<uf>[A-Z]{2})\.?$", text)
    if m:
        city = clean_cell(m.group("city"))
        uf = clean_cell(m.group("uf")).upper()
        return text, city, uf

    return text, "", ""


class SupplierDatabase:
    """
    Banco geral de fornecedores.

    Busca preferencial:
    1. CNPJ;
    2. razão social normalizada.

    Compatível com banco de mala direta:
    FORNECEDOR | ENDEREÇO | CEP | FONE | EMAIL | CNPJ | INSCRIÇÃO ESTAUDAL | REPRESENTANTE | CPF | RG | ORGAO
    """

    def __init__(self, path=None):
        self.path = Path(path) if path else None
        self.records = []
        self.by_cnpj = {}
        self.by_name = {}
        self.warnings = []

        if self.path and self.path.exists():
            self.load(self.path)

    @classmethod
    def empty(cls):
        return cls(None)

    def load(self, path):
        if openpyxl is None:
            self.warnings.append("openpyxl não disponível; banco de fornecedores não foi lido.")
            return

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        headers = [normalize_text(cell.value) for cell in ws[1]]
        col = self._map_headers(headers)

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_endereco = self._get(row, col, "endereco")
            endereco_limpo, municipio, uf = split_address_city_uf(raw_endereco)

            rec = {
                "fornecedor": self._get(row, col, "fornecedor"),
                "endereco": endereco_limpo or raw_endereco,
                "municipio": municipio,
                "uf": uf,
                "cep": self._get(row, col, "cep"),
                "telefone": self._get(row, col, "telefone"),
                "email": self._get(row, col, "email"),
                "cnpj": self._get(row, col, "cnpj"),
                "inscricao_estadual": self._get(row, col, "inscricao_estadual"),
                "representante": self._get(row, col, "representante"),
                "cpf_representante": self._get(row, col, "cpf"),
                "rg_representante": self._get(row, col, "rg"),
                "orgao_expedidor": self._get(row, col, "orgao"),
            }

            if not rec["fornecedor"] and not rec["cnpj"]:
                continue

            self.records.append(rec)

            cnpj_key = only_digits(rec["cnpj"])
            if cnpj_key:
                self.by_cnpj.setdefault(cnpj_key, []).append(rec)

            name_key = normalize_text(rec["fornecedor"])
            if name_key:
                self.by_name.setdefault(name_key, []).append(rec)

        self._register_duplicates()

    def _map_headers(self, headers):
        mapping = {}
        for idx, h in enumerate(headers):
            if "FORNECEDOR" in h or "RAZAO" in h:
                mapping["fornecedor"] = idx
            elif "ENDERE" in h:
                mapping["endereco"] = idx
            elif h == "CEP":
                mapping["cep"] = idx
            elif "FONE" in h or "TELEFONE" in h:
                mapping["telefone"] = idx
            elif "EMAIL" in h or "E-MAIL" in h:
                mapping["email"] = idx
            elif "CNPJ" in h:
                mapping["cnpj"] = idx
            elif "INSCRI" in h or "ESTAUDAL" in h or "ESTADUAL" in h:
                mapping["inscricao_estadual"] = idx
            elif "REPRESENT" in h:
                mapping["representante"] = idx
            elif h == "CPF" or "CPF" in h:
                mapping["cpf"] = idx
            elif h == "RG" or "IDENTIDADE" in h:
                mapping["rg"] = idx
            elif "ORGAO" in h or "ÓRGAO" in h or "EXPED" in h:
                mapping["orgao"] = idx
        return mapping

    def _get(self, row, col, key):
        idx = col.get(key)
        if idx is None or idx >= len(row):
            return ""
        return clean_cell(row[idx])

    def _register_duplicates(self):
        for cnpj, regs in self.by_cnpj.items():
            if len(regs) > 1:
                nomes = "; ".join(r.get("fornecedor", "") for r in regs)
                self.warnings.append(f"DUPLICIDADE NO BANCO: CNPJ {cnpj} aparece {len(regs)} vezes: {nomes}")

        for name, regs in self.by_name.items():
            if len(regs) > 1:
                cnpjs = "; ".join(r.get("cnpj", "") for r in regs)
                self.warnings.append(f"DUPLICIDADE NO BANCO: fornecedor '{regs[0].get('fornecedor','')}' aparece {len(regs)} vezes: {cnpjs}")

    def find(self, fornecedor_nome="", cnpj=""):
        cnpj_key = only_digits(cnpj)
        if cnpj_key and cnpj_key in self.by_cnpj:
            regs = self.by_cnpj[cnpj_key]
            if len(regs) > 1:
                return regs[0], f"Há mais de um fornecedor cadastrado para o CNPJ {cnpj}."
            return regs[0], ""

        name_key = normalize_text(fornecedor_nome)
        if name_key and name_key in self.by_name:
            regs = self.by_name[name_key]
            if len(regs) > 1:
                return regs[0], f"Há mais de um fornecedor cadastrado para o nome {fornecedor_nome}."
            return regs[0], ""

        if name_key:
            candidates = []
            for rec in self.records:
                rn = normalize_text(rec.get("fornecedor"))
                if name_key in rn or rn in name_key:
                    candidates.append(rec)
            if len(candidates) == 1:
                return candidates[0], ""
            if len(candidates) > 1:
                return candidates[0], f"Há mais de um fornecedor parecido cadastrado para {fornecedor_nome}."

        return None, ""

    def enrich_ata(self, ata):
        rec, warning = self.find(ata.fornecedor_nome, ata.fornecedor_cnpj)

        if warning:
            ata.inconsistencias.append(warning)

        if not rec:
            return ata

        fields = {
            "endereco": "endereco",
            "municipio": "municipio",
            "uf": "uf",
            "cep": "cep",
            "telefone": "telefone",
            "email": "email",
            "inscricao_estadual": "inscricao_estadual",
            "representante": "representante",
            "cpf_representante": "cpf_representante",
            "rg_representante": "rg_representante",
            "orgao_expedidor": "orgao_expedidor",
        }

        for ata_field, rec_field in fields.items():
            current = getattr(ata, ata_field, "")
            value = rec.get(rec_field, "")
            if (not current or "INFORMAÇÃO NÃO LOCALIZADA" in str(current).upper()) and value:
                setattr(ata, ata_field, value)

        if not getattr(ata, "fornecedor_cnpj", "") and rec.get("cnpj"):
            ata.fornecedor_cnpj = rec["cnpj"]

        return ata
