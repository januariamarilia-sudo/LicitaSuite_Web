"""
LicitaSuite - Módulo Opcional de Mala Direta de Fornecedores

Este módulo NÃO mexe em:
- tabelas;
- itens;
- valores;
- apêndice;
- cláusula 4;
- formatação das atas.

Ele serve apenas para ler um banco Excel de fornecedores
e preencher dados cadastrais quando o fornecedor for localizado.
"""

from pathlib import Path
import re
import unicodedata

try:
    import openpyxl
except Exception:
    openpyxl = None


def normalizar(texto):
    texto = "" if texto is None else str(texto)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return " ".join(texto.upper().split())


def digitos(texto):
    return re.sub(r"\D+", "", "" if texto is None else str(texto))


def limpar(texto):
    if texto is None:
        return ""
    texto = str(texto).strip()
    if texto.endswith(".0") and texto.replace(".0", "").isdigit():
        texto = texto[:-2]
    return " ".join(texto.split()).strip()


def separar_endereco_municipio_uf(endereco):
    """
    Aceita endereço em formato de mala direta:

    Rodovia JK 459, KM 99 S/N Galpão, Bairro Santa Edwirges,
    no Município de Pouso Alegre - MG
    """
    texto = limpar(endereco)
    if not texto:
        return "", "", ""

    padroes = [
        r"^(?P<end>.*?),?\s+no\s+Munic[ií]pio\s+de\s+(?P<mun>.+?)\s*[-/]\s*(?P<uf>[A-Z]{2})\.?$",
        r"^(?P<end>.*?),?\s+Munic[ií]pio\s+de\s+(?P<mun>.+?)\s*[-/]\s*(?P<uf>[A-Z]{2})\.?$",
        r"^(?P<end>.*?),?\s+em\s+(?P<mun>.+?)\s*[-/]\s*(?P<uf>[A-Z]{2})\.?$",
    ]

    for padrao in padroes:
        m = re.search(padrao, texto, flags=re.IGNORECASE)
        if m:
            return (
                limpar(m.group("end")).rstrip(","),
                limpar(m.group("mun")),
                limpar(m.group("uf")).upper(),
            )

    return texto, "", ""


class BancoMalaDiretaFornecedores:
    def __init__(self, arquivo_xlsx=None):
        self.arquivo_xlsx = Path(arquivo_xlsx) if arquivo_xlsx else None
        self.registros = []
        self.por_cnpj = {}
        self.por_nome = {}
        self.observacoes = []

        if self.arquivo_xlsx and self.arquivo_xlsx.exists():
            self.carregar(self.arquivo_xlsx)

    @classmethod
    def localizar_no_diretorio(cls, pasta):
        pasta = Path(pasta)
        candidatos = list(pasta.rglob("*.xlsx"))

        preferidos = []
        for arq in candidatos:
            nome = normalizar(arq.name)
            if "BANCO" in nome or "FORNECEDOR" in nome or "CADASTRO" in nome or "DADOS" in nome:
                preferidos.append(arq)

        if preferidos:
            return cls(preferidos[0])

        if candidatos:
            return cls(candidatos[0])

        return cls(None)

    def carregar(self, arquivo):
        if openpyxl is None:
            self.observacoes.append("Banco de fornecedores não lido: openpyxl não instalado.")
            return

        wb = openpyxl.load_workbook(arquivo, data_only=True)
        ws = wb.active

        cabecalhos = [normalizar(c.value) for c in ws[1]]
        col = self._mapear_colunas(cabecalhos)

        for row in ws.iter_rows(min_row=2, values_only=True):
            endereco_original = self._valor(row, col, "endereco")
            endereco, municipio, uf = separar_endereco_municipio_uf(endereco_original)

            registro = {
                "fornecedor": self._valor(row, col, "fornecedor"),
                "endereco": endereco or endereco_original,
                "municipio": municipio,
                "uf": uf,
                "cep": self._valor(row, col, "cep"),
                "telefone": self._valor(row, col, "telefone"),
                "email": self._valor(row, col, "email"),
                "cnpj": self._valor(row, col, "cnpj"),
                "inscricao_estadual": self._valor(row, col, "inscricao_estadual"),
                "representante": self._valor(row, col, "representante"),
                "cpf_representante": self._valor(row, col, "cpf"),
                "rg_representante": self._valor(row, col, "rg"),
                "orgao_expedidor": self._valor(row, col, "orgao"),
            }

            if not registro["fornecedor"] and not registro["cnpj"]:
                continue

            self.registros.append(registro)

            cnpj = digitos(registro["cnpj"])
            if cnpj:
                self.por_cnpj.setdefault(cnpj, []).append(registro)

            nome = normalizar(registro["fornecedor"])
            if nome:
                self.por_nome.setdefault(nome, []).append(registro)

        self._verificar_duplicidades()

    def _mapear_colunas(self, cabecalhos):
        mapa = {}
        for i, h in enumerate(cabecalhos):
            if "FORNECEDOR" in h or "RAZAO" in h:
                mapa["fornecedor"] = i
            elif "ENDERE" in h:
                mapa["endereco"] = i
            elif h == "CEP":
                mapa["cep"] = i
            elif "FONE" in h or "TELEFONE" in h:
                mapa["telefone"] = i
            elif "EMAIL" in h or "E-MAIL" in h:
                mapa["email"] = i
            elif "CNPJ" in h:
                mapa["cnpj"] = i
            elif "INSCRI" in h or "ESTAUDAL" in h or "ESTADUAL" in h:
                mapa["inscricao_estadual"] = i
            elif "REPRESENT" in h:
                mapa["representante"] = i
            elif h == "CPF" or "CPF" in h:
                mapa["cpf"] = i
            elif h == "RG" or "IDENTIDADE" in h:
                mapa["rg"] = i
            elif "ORGAO" in h or "EXPED" in h:
                mapa["orgao"] = i
        return mapa

    def _valor(self, row, col, chave):
        idx = col.get(chave)
        if idx is None or idx >= len(row):
            return ""
        return limpar(row[idx])

    def _verificar_duplicidades(self):
        for cnpj, regs in self.por_cnpj.items():
            if len(regs) > 1:
                nomes = "; ".join(r.get("fornecedor", "") for r in regs)
                self.observacoes.append(
                    f"Duplicidade no banco: CNPJ {cnpj} aparece {len(regs)} vezes: {nomes}"
                )

        for nome, regs in self.por_nome.items():
            if len(regs) > 1:
                cnpjs = "; ".join(r.get("cnpj", "") for r in regs)
                self.observacoes.append(
                    f"Duplicidade no banco: fornecedor {regs[0].get('fornecedor', '')} aparece {len(regs)} vezes: {cnpjs}"
                )

    def procurar(self, fornecedor_nome="", cnpj=""):
        cnpj_key = digitos(cnpj)

        if cnpj_key and cnpj_key in self.por_cnpj:
            regs = self.por_cnpj[cnpj_key]
            if len(regs) > 1:
                return regs[0], f"Há mais de um cadastro para o CNPJ {cnpj}."
            return regs[0], ""

        nome_key = normalizar(fornecedor_nome)
        if nome_key and nome_key in self.por_nome:
            regs = self.por_nome[nome_key]
            if len(regs) > 1:
                return regs[0], f"Há mais de um cadastro para o fornecedor {fornecedor_nome}."
            return regs[0], ""

        return None, ""

    def enriquecer_ata_data(self, ata):
        registro, obs = self.procurar(
            fornecedor_nome=getattr(ata, "fornecedor_nome", ""),
            cnpj=getattr(ata, "fornecedor_cnpj", ""),
        )

        if obs and hasattr(ata, "inconsistencias"):
            ata.inconsistencias.append(obs)

        if not registro:
            return ata

        campos = {
            "endereco": "endereco",
            "municipio": "municipio",
            "uf": "uf",
            "cep": "cep",
            "telefone": "telefone",
            "email": "email",
            "inscricao_estadual": "inscricao_estadual",
            "representante": "representante",
            "cpf_representante": "cpf_representante",
            "rg_representante": "rg_representante",
            "orgao_expedidor": "orgao_expedidor",
        }

        for campo_ata, campo_registro in campos.items():
            atual = getattr(ata, campo_ata, "")
            novo = registro.get(campo_registro, "")

            if novo and (not atual or "INFORMAÇÃO NÃO LOCALIZADA" in str(atual).upper()):
                setattr(ata, campo_ata, novo)

        return ata
