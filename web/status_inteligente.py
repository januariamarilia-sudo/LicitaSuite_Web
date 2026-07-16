from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED
import json
from typing import Any

try:
    from validador_vencedores import parse_vencedores_pdf, extract_pdf_from_zip
except Exception:
    parse_vencedores_pdf = None
    extract_pdf_from_zip = None


def load_report(output_dir: Path) -> dict[str, Any] | None:
    candidates = [
        output_dir / "relatorio_conferencia.json",
        output_dir.parent / "relatorio_conferencia.json",
        Path("output") / "relatorio_conferencia.json",
    ]

    for path in candidates:
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                return None

    return None


def build_status_data(result: Any, output_dir: Path, docx_files: list[Path], zip_path: Path | None = None) -> dict[str, Any]:
    messages = list(getattr(result, "messages", []) or [])
    errors = list(getattr(result, "errors", []) or [])
    report = load_report(output_dir)

    fornecedores = []
    itens_sem_vencedor = []
    observacoes_gerais = []

    if report:
        fornecedores = report.get("fornecedores", []) or report.get("atas", []) or []
        itens_sem_vencedor = report.get("itens_sem_vencedor", []) or report.get("itens_nao_localizados", []) or []
        observacoes_gerais = report.get("observacoes_gerais", []) or report.get("inconsistencias", []) or []

    # Fallback: quando o motor não entrega relatório, lê os fornecedores diretamente do PDF dentro do ZIP.
    if zip_path and extract_pdf_from_zip is not None and parse_vencedores_pdf is not None:
        try:
            pdf_path = extract_pdf_from_zip(zip_path)
            if pdf_path:
                fornecedores_pdf = parse_vencedores_pdf(pdf_path)

                # Usa PDF como base principal da planilha, pois ele contém todos os vencedores.
                if fornecedores_pdf:
                    fornecedores = fornecedores_pdf
        except Exception as exc:
            observacoes_gerais.append(f"Não foi possível montar prévia pelo PDF: {exc}")

    # Relaciona arquivos DOCX gerados quando possível.
    docx_names = [p.name for p in docx_files]

    for f in fornecedores:
        f.setdefault("observacoes", [])
        f.setdefault("arquivo_gerado", "")

        nome = (
            f.get("nome")
            or f.get("fornecedor")
            or f.get("contratado")
            or ""
        )

        if not f.get("arquivo_gerado"):
            nome_norm = "".join(ch for ch in nome.upper() if ch.isalnum())
            for docx_name in docx_names:
                docx_norm = "".join(ch for ch in docx_name.upper() if ch.isalnum())
                if nome_norm[:10] and nome_norm[:10] in docx_norm:
                    f["arquivo_gerado"] = docx_name
                    break

        if not f.get("arquivo_gerado"):
            obs = f.get("observacoes", [])
            if isinstance(obs, list):
                obs.append("ATA DOCX não localizada no ZIP gerado")
            else:
                f["observacoes"] = [str(obs), "ATA DOCX não localizada no ZIP gerado"]

    total_itens = 0
    for f in fornecedores:
        itens = f.get("itens", [])
        if isinstance(itens, list):
            total_itens += len(itens)

    faltantes = [
        f for f in fornecedores
        if not f.get("arquivo_gerado")
    ]

    if faltantes:
        observacoes_gerais.append(
            f"{len(faltantes)} fornecedor(es) do PDF sem ATA DOCX correspondente."
        )

    return {
        "messages": messages,
        "errors": errors,
        "fornecedores": fornecedores,
        "itens_sem_vencedor": itens_sem_vencedor,
        "observacoes_gerais": observacoes_gerais,
        "process_info": {},
        "total_atas": len(docx_files),
        "total_fornecedores_pdf": len(fornecedores),
        "total_itens": total_itens,
        "total_itens_nao_localizados": len(itens_sem_vencedor),
    }


def create_control_workbook(status_data: dict[str, Any], output_dir: Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except Exception as exc:
        raise RuntimeError("Para gerar a planilha de controle, inclua openpyxl no requirements.txt.") from exc

    output_dir.mkdir(parents=True, exist_ok=True)
    out = output_dir / "CONTROLE DE NUMERAÇÃO - PL ATUAL.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Controle"

    headers = [
        "Nº",
        "TIPO",
        "OBJETO",
        "PL",
        "MODALIDADE",
        "CONTRATADO",
        "CNPJ",
        "E-MAIL",
        "TELEFONE",
        "VALOR TOTAL",
        "QTDE ITENS",
        "ITENS",
        "ARQUIVO GERADO",
        "OBSERVAÇÃO",
    ]

    ws.append(headers)

    fill = PatternFill("solid", fgColor="073F9E")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fornecedores = status_data.get("fornecedores", [])

    for idx, f in enumerate(fornecedores, start=1):
        itens = f.get("itens", [])

        if isinstance(itens, list):
            itens_texto = ", ".join(str(i).zfill(4) for i in itens)
            qtde = len(itens)
        else:
            itens_texto = str(itens or "")
            qtde = ""

        obs = f.get("observacoes") or ""
        if isinstance(obs, list):
            obs = " | ".join(str(o) for o in obs if str(o).strip())

        ws.append([
            idx,
            "ATA",
            "",
            "",
            "",
            f.get("nome") or f.get("fornecedor") or f.get("contratado") or "",
            f.get("cnpj") or "",
            f.get("email") or f.get("e_mail") or "",
            f.get("telefone") or f.get("fone") or "",
            f.get("valor_total") or f.get("valor") or "",
            qtde,
            itens_texto,
            f.get("arquivo_gerado") or "",
            obs,
        ])

    widths = {
        "A": 8,
        "B": 12,
        "C": 34,
        "D": 16,
        "E": 18,
        "F": 44,
        "G": 22,
        "H": 32,
        "I": 22,
        "J": 20,
        "K": 13,
        "L": 36,
        "M": 58,
        "N": 50,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    resumo = wb.create_sheet("Resumo")
    resumo.append(["Resumo da Geração"])
    resumo["A1"].font = Font(bold=True, size=16, color="073F9E")
    resumo.append(["Total de atas DOCX geradas", status_data.get("total_atas", 0)])
    resumo.append(["Total de fornecedores no PDF", status_data.get("total_fornecedores_pdf", len(fornecedores))])
    resumo.append(["Total de itens identificados", status_data.get("total_itens", 0)])
    resumo.append(["Itens não localizados", ", ".join(str(i) for i in status_data.get("itens_sem_vencedor", []))])
    resumo.append(["Observações gerais", " | ".join(str(o) for o in status_data.get("observacoes_gerais", []))])

    resumo.column_dimensions["A"].width = 35
    resumo.column_dimensions["B"].width = 100

    for row in resumo.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(out)
    return out


def add_extra_files_to_zip(zip_path: Path, extra_files: list[Path]) -> Path:
    if not extra_files:
        return zip_path

    tmp = zip_path.with_suffix(".tmp.zip")

    with ZipFile(zip_path, "r") as zin:
        existing = {item.filename for item in zin.infolist()}

        with ZipFile(tmp, "w", ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                zout.writestr(item, zin.read(item.filename))

            for file in extra_files:
                if file.exists() and file.name not in existing:
                    zout.write(file, file.name)

    tmp.replace(zip_path)
    return zip_path
