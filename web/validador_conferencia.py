from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED
import json
import re
import xml.etree.ElementTree as ET
from typing import Any

import pdfplumber


SUPPLIER_HEADER_RE = re.compile(r"(?m)^\s*(.+?)\s*(?:\|\s*Tipo:|\s+-\s*Tipo:)", flags=re.I)
CNPJ_RE = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\s*\d{2}")
TOTAL_OFICIAL_RE = re.compile(
    r"(?:TOTAL\s+DO\s+VENCEDOR|Total)\s+R\$\s*(?P<total>\d{1,3}(?:\.\d{3})*,\d{2,4}|\d+,\d{2,4})",
    flags=re.I,
)
ITEM_START_RE = re.compile(r"(?m)^(\d{4})\s+")
MONEY_RE = re.compile(r"R\$\s*(\d{1,3}(?:\.\d{3})*,\d{2,4}|\d+,\d{2,4})")
ANY_MONEY_RE = re.compile(r"(?<!\d)(\d{1,3}(?:\.\d{3})*,\d{2,4}|\d+,\d{2,4})(?!\d)")


def _money(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("R$", "").strip()
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except Exception:
        return 0.0


def _fmt_money(value: Any) -> str:
    value = _money(value)
    return "R$ " + f"{value:,.4f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _norm(value: Any) -> str:
    text = str(value or "").upper()
    for a, b in {
        "Á": "A", "À": "A", "Â": "A", "Ã": "A",
        "É": "E", "Ê": "E", "Í": "I",
        "Ó": "O", "Ô": "O", "Õ": "O",
        "Ú": "U", "Ç": "C",
    }.items():
        text = text.replace(a, b)
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _supplier_key(name: Any) -> str:
    text = _norm(name)
    stopwords = {
        "LTDA", "EIRELI", "SA", "S", "A", "ME", "EPP", "SS",
        "COMERCIO", "COMERCIAL", "DISTRIBUIDORA", "DISTRIBUICAO",
        "MEDICAMENTOS", "PRODUTOS", "PRODUTO", "HOSPITALARES",
        "IMPORTACAO", "EXPORTACAO", "SERVICOS", "FARMACEUTICA",
        "FARMACEUTICOS",
    }
    tokens = [t for t in text.split() if t not in stopwords]
    return " ".join(tokens[:5])


def _short_key(name: Any) -> str:
    return " ".join(_supplier_key(name).split()[:2])


def _load_engine_report_rows(output_dir: Path) -> list[dict[str, Any]]:
    candidates = [
        output_dir / "relatorio_conferencia.json",
        output_dir.parent / "relatorio_conferencia.json",
        Path("output") / "relatorio_conferencia.json",
    ]

    data = None
    for path in candidates:
        if not path.exists():
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            break
        except Exception:
            data = None

    if not data:
        return []

    rows = data.get("fornecedores", []) or data.get("atas", []) or []
    normalized = []
    for row in rows:
        nome = row.get("nome") or row.get("fornecedor") or row.get("contratado") or ""
        normalized.append({
            **row,
            "nome": nome,
            "key": _supplier_key(nome),
            "short_key": _short_key(nome),
        })
    return normalized


def _find_matching_row(base: dict[str, Any], candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    key = base.get("key", "")
    short = base.get("short_key", "")
    for row in candidates:
        rk = row.get("key", "")
        rs = row.get("short_key", "")
        key_first = key.split()[0] if key.split() else ""
        rk_first = rk.split()[0] if rk.split() else ""
        if key and rk and key == rk:
            return row
        if short and rs and short == rs:
            return row
        if key_first and rk_first and len(key_first) >= 3 and key_first == rk_first:
            return row
        if key and rk and min(len(key), len(rk)) >= 6 and (key in rk or rk in key):
            return row
        if short and rk and len(short) >= 6 and short in rk:
            return row
    return None


def _prepare_text(text: str) -> str:
    text = text or ""
    text = text.replace("\xa0", " ")
    text = text.replace("\u00ad", "")
    text = text.replace("\ufffe", "-")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text


def _read_pdf_text(pdf_path: Path) -> str:
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            pages.append(page.extract_text(layout=True, x_tolerance=1, y_tolerance=3) or "")
    return "\n".join(pages)


def _find_pdf_in_zip_or_folder(path: Path | None) -> Path | None:
    if not path:
        return None

    if path.is_file() and path.suffix.lower() == ".pdf":
        return path

    if path.is_dir():
        pdfs = list(path.rglob("*.pdf"))
        if not pdfs:
            return None
        pdfs.sort(key=lambda p: (0 if "venced" in p.name.lower() or "itensvencidos" in p.name.lower() else 1, len(p.name)))
        return pdfs[0]

    if path.is_file() and path.suffix.lower() == ".zip":
        temp = Path("temp/conferencia_pdf")
        temp.mkdir(parents=True, exist_ok=True)

        with ZipFile(path, "r") as z:
            pdfs = [n for n in z.namelist() if n.lower().endswith(".pdf")]
            if not pdfs:
                return None
            pdfs.sort(key=lambda n: (0 if "venced" in n.lower() or "itensvencidos" in n.lower() else 1, len(n)))
            out = temp / Path(pdfs[0]).name
            out.write_bytes(z.read(pdfs[0]))
            return out

    return None


def _split_blocks(text: str) -> list[str]:
    matches = []
    for m in SUPPLIER_HEADER_RE.finditer(text):
        window = text[m.start():m.start() + 1800]
        if CNPJ_RE.search(window):
            matches.append(m)

    blocks = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        blocks.append(text[start:end].strip())
    return blocks


def _table_region(block: str) -> str:
    header = re.search(r"C[oó]digo\s+Produto.*?Valor\s+Total", block, flags=re.I | re.S)
    if not header:
        return ""

    start = header.end()

    stops = []
    m1 = re.search(r"TOTAL\s+DO\s+VENCEDOR", block[start:], flags=re.I)
    if m1:
        stops.append(start + m1.start())

    m2 = re.search(r"(?m)^\s*Total\s+R\$", block[start:], flags=re.I)
    if m2:
        stops.append(start + m2.start())

    end = min(stops) if stops else len(block)
    return block[start:end].strip()


def _parse_pdf_financeiro(pdf_path: Path) -> list[dict[str, Any]]:
    text = _prepare_text(_read_pdf_text(pdf_path))
    rows = []

    for block in _split_blocks(text):
        name_m = SUPPLIER_HEADER_RE.search(block)
        if not name_m:
            continue

        nome = re.sub(r"\s+", " ", name_m.group(1)).strip(" -|")
        totals = [_money(m.group("total")) for m in TOTAL_OFICIAL_RE.finditer(block)]
        total_oficial = totals[-1] if totals else 0.0

        table = _table_region(block)
        item_values = []
        item_codes = []

        if table:
            starts = list(ITEM_START_RE.finditer(table))
            for i, s in enumerate(starts):
                item_codes.append(s.group(1))
                start = s.start()
                end = starts[i + 1].start() if i + 1 < len(starts) else len(table)
                row = re.sub(r"\s+", " ", table[start:end]).strip()

                vals = MONEY_RE.findall(row)
                if len(vals) >= 2:
                    item_values.append(_money(vals[-1]))
                elif vals:
                    after = row[row.find(vals[-1]) + len(vals[-1]):]
                    loose = [x for x in ANY_MONEY_RE.findall(after) if x != vals[-1]]
                    item_values.append(_money(loose[-1]) if loose else 0.0)
                else:
                    nums = ANY_MONEY_RE.findall(row)
                    item_values.append(_money(nums[-1]) if nums else 0.0)

        soma_itens = sum(item_values)

        rows.append({
            "nome": nome,
            "key": _supplier_key(nome),
            "short_key": _short_key(nome),
            "itens_pdf": item_codes,
            "total_oficial_vencedor": total_oficial,
            "soma_itens_pdf": soma_itens,
            "status_soma_pdf": "OK" if total_oficial and abs(soma_itens - total_oficial) <= 0.05 else "DIVERGENTE - CORRIGIR MANUALMENTE",
        })

    return rows


def _docx_text(path: Path) -> str:
    try:
        with ZipFile(path, "r") as z:
            xml_names = [n for n in z.namelist() if n.startswith("word/") and n.endswith(".xml")]
            parts = []
            ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
            for name in xml_names:
                if not (name.endswith("document.xml") or "header" in name or "footer" in name):
                    continue
                root = ET.fromstring(z.read(name))
                for node in root.findall(".//w:t", ns):
                    if node.text:
                        parts.append(node.text)
            return " ".join(parts)
    except Exception:
        return ""


def _docx_rows(docx_files: list[Path]) -> list[dict[str, Any]]:
    rows = []
    for file in docx_files:
        stem = file.stem
        nome = stem.split(" - ", 1)[1] if " - " in stem else stem
        text = _docx_text(file)
        rows.append({
            "arquivo": file.name,
            "nome": nome,
            "key": _supplier_key(nome),
            "short_key": _short_key(nome),
            "money_values": [_money(v) for v in MONEY_RE.findall(text)],
            "email": _extract_email(text),
            "telefone": _extract_telefone(text),
        })
    return rows


def _extract_email(text: str) -> str:
    match = re.search(r"[\w.+-]+@[\w.-]+\.\w+", text or "")
    return match.group(0) if match else ""


def _extract_telefone(text: str) -> str:
    match = re.search(r"(?:telefone|fone)\s*[:\-]?\s*([()0-9 .+\-]{8,24})", text or "", flags=re.I)
    return re.sub(r"\s+", " ", match.group(1)).strip(" .-") if match else ""


def _match_rows(pdf_rows, docx_rows):
    result = {}
    for pdf in pdf_rows:
        key = pdf.get("key", "")
        short = pdf.get("short_key", "")
        found = None

        for docx in docx_rows:
            dk = docx.get("key", "")
            ds = docx.get("short_key", "")
            key_first = key.split()[0] if key.split() else ""
            dk_first = dk.split()[0] if dk.split() else ""
            if key and dk and key == dk:
                found = docx
                break
            if short and ds and short == ds:
                found = docx
                break
            if key_first and dk_first and len(key_first) >= 3 and key_first == dk_first:
                found = docx
                break
            if key and dk and min(len(key), len(dk)) >= 6 and (key in dk or dk in key):
                found = docx
                break
            if short and dk and len(short) >= 6 and short in dk:
                found = docx
                break

        if found:
            result[key] = found

    return result


def build_conferencia(zip_path: Path | None, output_dir: Path, docx_files: list[Path]) -> dict[str, Any]:
    pdf_path = _find_pdf_in_zip_or_folder(zip_path)
    pdf_rows = _parse_pdf_financeiro(pdf_path) if pdf_path else []
    report_rows = _load_engine_report_rows(output_dir)
    docx_rows = _docx_rows(docx_files)

    matched = _match_rows(pdf_rows, docx_rows)

    linhas = []
    faltando_docx = []
    base_rows = report_rows or pdf_rows

    for pdf in base_rows:
        pdf_ref = _find_matching_row(pdf, pdf_rows) if report_rows else pdf
        key = pdf_ref.get("key", "") if pdf_ref else pdf.get("key", "")
        docx = matched.get(key) or _find_matching_row(pdf, docx_rows)
        total_oficial = (
            pdf.get("valor_total") or pdf.get("valor")
            if report_rows and (pdf.get("valor_total") or pdf.get("valor"))
            else pdf_ref.get("total_oficial_vencedor") if pdf_ref and pdf_ref.get("total_oficial_vencedor") else 0
        )
        total_oficial = _money(total_oficial)
        status_ata = "DOCX NÃO LOCALIZADO"
        valor_na_ata = ""
        obs = ""

        if not docx:
            faltando_docx.append(pdf["nome"])
        else:
            exists = any(abs(v - total_oficial) <= 0.05 for v in docx.get("money_values", []))
            if exists:
                status_ata = "OK"
                valor_na_ata = _fmt_money(total_oficial)
            else:
                status_ata = "DIVERGENTE - TOTAL OFICIAL NÃO LOCALIZADO NA ATA"
                valor_na_ata = "NÃO LOCALIZADO"
                obs = "Conferir manualmente. O TOTAL DO VENCEDOR/Total R$ não foi localizado na ata."

        linhas.append({
            "fornecedor": pdf["nome"],
            "itens_pdf": ", ".join(str(i).zfill(4) for i in (pdf.get("itens") or (pdf_ref.get("itens_pdf", []) if pdf_ref else []))),
            "total_vencedor_pdf": total_oficial,
            "soma_itens_pdf": pdf_ref.get("soma_itens_pdf", total_oficial) if pdf_ref else total_oficial,
            "status_soma_pdf": pdf_ref.get("status_soma_pdf", "OK") if pdf_ref else "OK",
            "email": pdf.get("email") or pdf.get("e_mail") or (docx.get("email", "") if docx else ""),
            "telefone": pdf.get("telefone") or pdf.get("fone") or (docx.get("telefone", "") if docx else ""),
            "valor_na_ata": valor_na_ata,
            "status_ata": status_ata,
            "observacao": obs,
        })

    divergencias = [x for x in linhas if x["status_ata"] != "OK"]

    return {
        "pdf_path": str(pdf_path) if pdf_path else "",
        "total_fornecedores_pdf": len(base_rows) if base_rows else len(pdf_rows),
        "total_docx": len(docx_files),
        "valor_total_pdf_oficial": sum(x["total_vencedor_pdf"] for x in linhas),
        "soma_itens_pdf": sum(x["soma_itens_pdf"] for x in linhas),
        "faltando_docx": faltando_docx,
        "linhas_financeiras": linhas,
        "divergencias_financeiras": divergencias,
        "ok": not faltando_docx and not divergencias,
    }


def write_conferencia_xlsx(conferencia: dict[str, Any], output_dir: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    output_dir.mkdir(parents=True, exist_ok=True)
    out = output_dir / "RELATORIO_CONFERENCIA_AUTOMATICA.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    rows = [
        ["Conferência financeira oficial", ""],
        ["PDF usado", conferencia.get("pdf_path", "")],
        ["Fornecedores no PDF", conferencia.get("total_fornecedores_pdf", 0)],
        ["Arquivos DOCX gerados", conferencia.get("total_docx", 0)],
        ["Valor total", _fmt_money(conferencia.get("valor_total_pdf_oficial", 0))],
        ["Status", "OK" if conferencia.get("ok") else "VERIFICAR PENDÊNCIAS"],
    ]

    for row in rows:
        ws.append(row)

    red = PatternFill("solid", fgColor="F8D7DA")
    green = PatternFill("solid", fgColor="D9EAD3")
    blue = PatternFill("solid", fgColor="073F9E")
    white = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"].font = Font(bold=True, size=15, color="073F9E")
    ws["B6"].fill = green if conferencia.get("ok") else red
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 100

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    sh = wb.create_sheet("Conferência Financeira")
    headers = [
        "FORNECEDOR",
        "E-MAIL",
        "TELEFONE",
        "ITENS PDF",
        "VALOR TOTAL",
        "VALOR LOCALIZADO NA ATA",
        "STATUS ATA",
        "OBSERVAÇÃO / CORREÇÃO MANUAL",
    ]
    sh.append(headers)

    for cell in sh[1]:
        cell.fill = blue
        cell.font = white
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in conferencia.get("linhas_financeiras", []):
        sh.append([
            row["fornecedor"],
            row.get("email", ""),
            row.get("telefone", ""),
            row["itens_pdf"],
            _fmt_money(row["total_vencedor_pdf"]),
            row["valor_na_ata"],
            row["status_ata"],
            row["observacao"],
        ])

    for row in sh.iter_rows(min_row=2):
        fill = red if row[6].value != "OK" else None
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if fill:
                cell.fill = fill

    widths = [42, 32, 22, 28, 24, 26, 42, 70]
    for idx, width in enumerate(widths, start=1):
        sh.column_dimensions[chr(64 + idx)].width = width

    sh.freeze_panes = "A2"
    sh.auto_filter.ref = sh.dimensions

    wb.save(out)
    return out


def add_conferencia_to_zip(zip_path: Path, conferencia_file: Path) -> Path:
    if not conferencia_file or not conferencia_file.exists():
        return zip_path

    tmp = zip_path.with_suffix(".tmp.zip")

    with ZipFile(zip_path, "r") as zin:
        existing = {item.filename for item in zin.infolist()}
        with ZipFile(tmp, "w", ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                zout.writestr(item, zin.read(item.filename))
            if conferencia_file.name not in existing:
                zout.write(conferencia_file, conferencia_file.name)

    tmp.replace(zip_path)
    return zip_path


def format_conferencia_markdown(conferencia: dict[str, Any]) -> str:
    if conferencia.get("ok"):
        return (
            "✅ Conferência financeira oficial OK\n\n"
            f"- Fornecedores no PDF: **{conferencia.get('total_fornecedores_pdf', 0)}**\n"
            f"- DOCX gerados: **{conferencia.get('total_docx', 0)}**\n"
            f"- Valor total oficial PDF: **{_fmt_money(conferencia.get('valor_total_pdf_oficial', 0))}**"
        )

    parts = [
        "⚠️ Conferência financeira oficial encontrou pendências",
        "",
        f"- Fornecedores no PDF: **{conferencia.get('total_fornecedores_pdf', 0)}**",
        f"- DOCX gerados: **{conferencia.get('total_docx', 0)}**",
        f"- Valor total oficial PDF: **{_fmt_money(conferencia.get('valor_total_pdf_oficial', 0))}**",
        "",
        "**Pendências principais:**",
    ]

    for row in conferencia.get("divergencias_financeiras", [])[:12]:
        parts.append(f"- {row['fornecedor']}: {row['status_ata']}")

    return "\n".join(parts)
